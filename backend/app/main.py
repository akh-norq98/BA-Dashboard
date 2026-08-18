from datetime import date, datetime, timedelta
import csv
import io
from fastapi import Depends, FastAPI, HTTPException, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from sqlalchemy import func, inspect, select, text
from sqlalchemy.orm import Session
from .config import settings
from .database import Base, SessionLocal, engine, get_db
from .models import ActionItem, ActionItemHistory, Bug, Client, DiscussionPoint, Feature, Meeting, Project, ProjectModule, QASummary, Requirement, Role
from .models import User, ProjectMember
from .schemas import ActionCreate, ActionOut, ActionUpdate, AdminUserCreate, AuthUser, BugCreate, BugOut, ClientOut, DiscussionPointCreate, DiscussionPointOut, DiscussionPointUpdate, FeatureCreate, FeatureOut, FeatureUpdate, LoginRequest, MeetingCreate, MeetingOut, MeetingUpdate, QASummaryCreate, QASummaryOut, RegisterRequest, RequirementCreate, RequirementOut, RequirementUpdate, TokenResponse
from .security import create_access_token, get_current_user, hash_password, verify_password

app = FastAPI(title="Delivery Hub API")
app.add_middleware(CORSMiddleware, allow_origins=settings.cors_origin_list, allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


@app.get("/health")
def health():
    return {"status": "ok"}

RBAC_MODULES = ["Action Items", "Meetings", "Discussions", "Reports", "Users", "Projects", "Clients", "Modules", "Dashboards"]
RBAC_ACTIONS = ["create", "read", "update", "delete", "view"]
def permissions_for(role: str) -> dict:
    if role == "admin": return {module: {action: True for action in RBAC_ACTIONS} for module in RBAC_MODULES}
    if role == "manager": return {module: {action: action != "delete" or module not in {"Users", "Clients"} for action in RBAC_ACTIONS} for module in RBAC_MODULES}
    if role == "editor": return {module: {action: action in {"create", "read", "update", "view"} for action in RBAC_ACTIONS} for module in RBAC_MODULES}
    return {module: {action: action in {"read", "view"} for action in RBAC_ACTIONS} for module in RBAC_MODULES}


@app.post("/api/auth/register", response_model=TokenResponse, status_code=201)
def register(payload: RegisterRequest, db: Session = Depends(get_db)):
    if db.query(User).filter(User.email == payload.email.lower()).first(): raise HTTPException(409, "Email is already registered")
    if payload.organization_id is not None:
        raise HTTPException(403, "Self-registration cannot select an organization")
    role = "viewer"
    role_record = db.query(Role).filter(Role.name == role).first()
    user = User(name=payload.name, email=payload.email.lower(), password_hash=hash_password(payload.password), organization_id=payload.organization_id, role=role, role_id=role_record.id if role_record else None)
    db.add(user); db.commit(); db.refresh(user)
    return {"access_token": create_access_token(user), "user": user}


@app.post("/api/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == payload.email.lower()).first()
    if not user or not verify_password(payload.password, user.password_hash): raise HTTPException(401, "Invalid email or password")
    return {"access_token": create_access_token(user), "user": user}


@app.get("/api/auth/me", response_model=AuthUser)
def me(user: User = Depends(get_current_user)):
    return user


@app.get("/api/me/projects")
def my_projects(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    projects = db.query(Project).filter(Project.organization_id == user.organization_id).all()
    return [{"id": project.id, "name": project.name, "status": project.status, "role": user.role} for project in projects]


@app.get("/api/auth/me/permissions")
def my_permissions(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    role = db.get(Role, user.role_id) if user.role_id else None
    return {"role": user.role, "permissions": role.permissions if role else permissions_for(user.role)}


def require_admin(user: User):
    if user.role != "admin": raise HTTPException(403, "Administrator access required")


def require_write(user: User):
    if user.role not in {"admin", "manager", "editor"}:
        raise HTTPException(403, "Write access required")


def require_delete(user: User):
    if user.role not in {"admin", "manager"}:
        raise HTTPException(403, "Delete access required")


def can_access_client(user: User, client_id: int | None) -> bool:
    return user.role == "admin" or client_id == user.organization_id


def require_client_scope(client_id: int | None, user: User):
    if client_id is None or not can_access_client(user, client_id):
        raise HTTPException(403, "Organization scope violation")


def scoped_record(item, user: User, message: str):
    if not item or not can_access_client(user, item.client_id):
        raise HTTPException(404, message)
    return item


@app.get("/api/roles")
def list_roles(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_admin(user)
    return [{"id": role.id, "name": role.name, "description": role.description, "is_system": role.is_system, "permissions": role.permissions or {}} for role in db.query(Role).order_by(Role.name).all()]


@app.post("/api/roles", status_code=201)
def create_role(payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_admin(user)
    if not payload.get("name"): raise HTTPException(422, "Role name is required")
    if db.query(Role).filter(Role.name == payload["name"]).first(): raise HTTPException(409, "Role already exists")
    role = Role(name=payload["name"], description=payload.get("description", ""), permissions=payload.get("permissions", {}), is_system=False); db.add(role); db.commit(); db.refresh(role)
    return {"id": role.id, "name": role.name, "description": role.description, "is_system": role.is_system, "permissions": role.permissions}


@app.patch("/api/roles/{role_id}")
def update_role(role_id: int, payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_admin(user); role = db.get(Role, role_id)
    if not role: raise HTTPException(404, "Role not found")
    if role.is_system and payload.get("name") and payload["name"] != role.name: raise HTTPException(400, "System role names cannot change")
    if "description" in payload: role.description = payload["description"]
    if "permissions" in payload: role.permissions = payload["permissions"]
    db.commit(); db.refresh(role); return {"id": role.id, "name": role.name, "description": role.description, "is_system": role.is_system, "permissions": role.permissions}


@app.delete("/api/roles/{role_id}", status_code=204)
def delete_role(role_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_admin(user); role = db.get(Role, role_id)
    if not role: raise HTTPException(404, "Role not found")
    if role.is_system: raise HTTPException(400, "System roles cannot be deleted")
    if db.query(User).filter(User.role_id == role_id).count(): raise HTTPException(400, "Role is assigned to users")
    db.delete(role); db.commit()


@app.get("/api/users")
def users(project_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if project_id is not None:
        scoped_record(db.get(Project, project_id), user, "Project not found")
    query = db.query(User).filter(User.is_active.is_(True)) if user.role == "admin" and not project_id else db.query(User).filter(User.organization_id == user.organization_id, User.is_active.is_(True))
    if project_id:
        query = query.join(ProjectMember, ProjectMember.user_id == User.id).filter(ProjectMember.project_id == project_id)
    return [{"id": item.id, "name": item.name, "email": item.email, "role": item.role, "role_id": item.role_id} for item in query.order_by(User.name).all()]


@app.post("/api/users", status_code=201)
def create_user(payload: AdminUserCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_admin(user)
    email = payload.email.strip().lower()
    if payload.password != payload.password_confirmation:
        raise HTTPException(422, "Passwords do not match")
    if db.query(User).filter(User.email == email).first():
        raise HTTPException(409, "Email is already registered")
    client = db.get(Client, payload.organization_id)
    role = db.get(Role, payload.role_id)
    if not client: raise HTTPException(404, "Client organization not found")
    if not role: raise HTTPException(404, "Role not found")
    created = User(name=payload.name.strip(), email=email, password_hash=hash_password(payload.password), organization_id=client.id, role=role.name, role_id=role.id)
    db.add(created); db.commit(); db.refresh(created)
    return {"id": created.id, "name": created.name, "email": created.email, "organization_id": created.organization_id, "role": created.role, "role_id": created.role_id, "is_active": created.is_active}


@app.patch("/api/users/{user_id}/role")
def assign_role(user_id: int, payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_admin(user)
    target = db.get(User, user_id); role = db.get(Role, payload.get("role_id"))
    if not target or not role: raise HTTPException(404, "User or role not found")
    target.role_id = role.id; target.role = role.name; db.commit(); db.refresh(target)
    return {"id": target.id, "name": target.name, "email": target.email, "role": target.role, "role_id": target.role_id}


def scoped_action(action_id: int, user: User, db: Session) -> ActionItem:
    item = db.get(ActionItem, action_id)
    if not item or not can_access_client(user, item.client_id):
        raise HTTPException(404, "Action item not found")
    return item


@app.get("/api/action-items", response_model=list[ActionOut])
def action_items(assigned_to: int | None = None, status: str | None = None, priority: str | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    query = db.query(ActionItem)
    if user.role != "admin":
        query = query.filter(ActionItem.client_id == user.organization_id)
    if assigned_to is not None: query = query.filter(ActionItem.assigned_to == assigned_to)
    if status: query = query.filter(ActionItem.status == status)
    if priority: query = query.filter(ActionItem.priority == priority)
    return query.order_by(ActionItem.created_at.desc()).all()


@app.post("/api/action-items", response_model=ActionOut, status_code=201)
def create_action_item(payload: ActionCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user)
    require_client_scope(payload.client_id, user)
    values = payload.model_dump(exclude_unset=True)
    values["created_by"] = user.id
    values["status"] = payload.status or "Open"
    if values.get("due_date") is None:
        values["due_date"] = date.today()
    item = ActionItem(**values)
    db.add(item); db.commit(); db.refresh(item); return item


@app.patch("/api/action-items/{action_id}", response_model=ActionOut)
def update_action_item(action_id: int, payload: ActionUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    item = scoped_action(action_id, user, db)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    db.add(ActionItemHistory(action_item_id=item.id, changed_by=user.id, field_name="updated", old_value=None, new_value=item.status))
    db.commit(); db.refresh(item); return item


@app.get("/api/dashboards/personal")
def personal_dashboard(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Return the authenticated user's organization-scoped delivery snapshot."""
    project_ids = [project.id for project in db.query(Project).filter(Project.organization_id == user.organization_id).all()]
    client_id = user.organization_id
    actions_query = db.query(ActionItem).filter(ActionItem.client_id == client_id)
    if project_ids:
        actions_query = actions_query.filter((ActionItem.project_id.in_(project_ids)) | (ActionItem.project_id.is_(None)))
    actions = actions_query.all()
    today = date.today()

    def status_is(item, *values):
        return (item.status or "").strip().lower() in {value.lower() for value in values}

    open_items = [item for item in actions if not status_is(item, "Done", "Completed", "Closed")]
    overdue = [item for item in open_items if (item.eda or item.due_date) and (item.eda or item.due_date) < today]
    on_track = [item for item in open_items if item not in overdue and status_is(item, "Open", "To do", "In Progress", "On Track")]
    at_risk = [item for item in open_items if status_is(item, "At Risk")]
    completed = [item for item in actions if status_is(item, "Done", "Completed", "Closed")]

    modules = {module.id: module.name for module in db.query(ProjectModule).filter(ProjectModule.project_id.in_(project_ids)).all()} if project_ids else {}
    module_counts = {}
    for item in actions:
        name = modules.get(item.module_id, "Unassigned")
        module_counts[name] = module_counts.get(name, 0) + 1
    recent = sorted(actions, key=lambda item: item.created_at or datetime.min, reverse=True)[:5]
    upcoming = sorted([item for item in actions if item.eda and today <= item.eda <= today + timedelta(days=7)], key=lambda item: item.eda)[:10]

    recent_notes = db.query(DiscussionPoint).filter(DiscussionPoint.client_id == client_id).order_by(DiscussionPoint.created_at.desc()).limit(5).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).all()
    now = datetime.utcnow()
    current_window = now - timedelta(days=14)
    previous_window = now - timedelta(days=28)
    current_count = sum(1 for item in actions if item.created_at and item.created_at >= current_window)
    previous_count = sum(1 for item in actions if item.created_at and previous_window <= item.created_at < current_window)
    change = round(((current_count - previous_count) / previous_count) * 100) if previous_count else (100 if current_count else 0)

    return {
        "summary": {"total_action_items": len(actions), "total_change": change, "on_track": len(on_track), "at_risk": len(at_risk), "overdue": len(overdue), "completed": len(completed)},
        "module_breakdown": [{"name": name, "count": count} for name, count in sorted(module_counts.items(), key=lambda pair: pair[1], reverse=True)],
        "eda_performance": {
            "on_or_before": sum(1 for item in completed if item.eda and item.ada and item.ada <= item.eda),
            "completed_after": sum(1 for item in completed if item.eda and item.ada and item.ada > item.eda),
            "not_completed": len(open_items),
        },
        "recent_items": [{"id": item.id, "title": item.title, "module": modules.get(item.module_id, "Unassigned"), "status": item.status} for item in recent],
        "upcoming_deadlines": [{"title": item.title, "module": modules.get(item.module_id, "Unassigned"), "eda": item.eda.isoformat()} for item in upcoming],
        "discussion_summary": {"total": db.query(DiscussionPoint).filter(DiscussionPoint.client_id == client_id).count(), "recent": [{"id": point.id, "title": point.title, "summary": point.content, "status": point.status} for point in recent_notes]},
        "meeting_summary": {"total": len(meetings), "upcoming": sum(1 for meeting in meetings if meeting.scheduled_for > now), "completed": sum(1 for meeting in meetings if meeting.scheduled_for <= now)},
    }


@app.get("/api/dashboards/client")
def client_dashboard(project_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id, Project.organization_id == user.organization_id).first() if project_id else db.query(Project).filter(Project.organization_id == user.organization_id).first()
    if not project: raise HTTPException(404, "Project not found")
    actions = db.query(ActionItem).filter(ActionItem.client_id == user.organization_id).filter((ActionItem.project_id == project.id) | (ActionItem.project_id.is_(None))).all()
    completed_statuses = {"done", "completed", "closed"}
    completed = [item for item in actions if (item.status or "").lower() in completed_statuses]
    today = date.today()
    upcoming = sorted([item for item in actions if (item.status or "").lower() not in completed_statuses and item.eda and today <= item.eda <= today + timedelta(days=14)], key=lambda item: item.eda)[:8]
    recent = sorted([item for item in completed if item.created_at and item.created_at >= datetime.utcnow() - timedelta(days=30)], key=lambda item: item.created_at, reverse=True)[:8]
    pending = [item for item in actions if "pending_approval" in (item.tags or [])]
    meetings = db.query(Meeting).filter(Meeting.client_id == user.organization_id).all()
    now = datetime.utcnow(); next_meeting = min((meeting for meeting in meetings if meeting.scheduled_for >= now), key=lambda meeting: meeting.scheduled_for, default=None)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return {"project": {"id": project.id, "name": project.name, "client_name": db.get(Client, user.organization_id).name if user.organization_id else "Client"}, "progress": round(len(completed) / len(actions) * 100) if actions else 0, "upcoming_deliverables": [{"id": item.id, "title": item.title, "eda": item.eda.isoformat()} for item in upcoming], "recent_deliveries": [{"id": item.id, "title": item.title, "completed_at": item.created_at.isoformat()} for item in recent], "pending_approvals": [{"id": item.id, "title": item.title} for item in pending], "meetings": {"next": {"id": next_meeting.id, "title": next_meeting.title, "scheduled_at": next_meeting.scheduled_for.isoformat()} if next_meeting else None, "this_month": sum(1 for meeting in meetings if meeting.scheduled_for >= month_start)}, "total_tasks": len(actions), "completed_tasks": len(completed)}


@app.get("/api/dashboards/executive")
def executive_dashboard(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}: raise HTTPException(403, "Executive dashboard access is restricted")
    projects = db.query(Project).all() if user.role == "admin" else db.query(Project).filter(Project.organization_id == user.organization_id).all()
    project_ids = [project.id for project in projects]
    if not project_ids:
        return {"summary": {"total_projects": 0, "total_clients": 0, "total_action_items": 0, "total_completed": 0, "overall_progress": 0, "overdue": 0, "blocked": 0}, "project_health": [], "module_heatmap": [], "team_workload": [], "overdue_by_module": [], "recent_activity": []}
    actions = db.query(ActionItem).filter(ActionItem.project_id.in_(project_ids)).all()
    completed_statuses = {"done", "completed", "closed"}; today = date.today()
    completed = [item for item in actions if (item.status or "").lower() in completed_statuses]
    overdue = [item for item in actions if (item.status or "").lower() not in completed_statuses and item.eda and item.eda < today]
    blocked = [item for item in actions if (item.status or "").lower() == "blocked"]
    clients_by_id = {client.id: client.name for client in db.query(Client).all()}
    modules_by_id = {module.id: module.name for module in db.query(ProjectModule).all()}
    users_by_id = {member.id: member.name for member in db.query(User).all()}
    health = []
    for project in projects:
        items = [item for item in actions if item.project_id == project.id]; done = [item for item in items if (item.status or "").lower() in completed_statuses]; late = [item for item in items if (item.status or "").lower() not in completed_statuses and item.eda and item.eda < today]; blocked_count = sum((item.status or "").lower() == "blocked" for item in items)
        health_status = "critical" if len(late) > 5 or blocked_count > 3 else "warning" if len(late) > 2 or blocked_count > 1 else "healthy"
        health.append({"project_id": project.id, "project_name": project.name, "client_name": clients_by_id.get(project.organization_id, "Unknown"), "total_tasks": len(items), "completed_tasks": len(done), "progress": round(len(done) / len(items) * 100) if items else 0, "overdue": len(late), "blocked": blocked_count, "health": health_status})
    order = {"critical": 0, "warning": 1, "healthy": 2}; health.sort(key=lambda item: order[item["health"]])
    module_counts = {}; workload = {}; overdue_modules = {}
    for item in actions:
        module_name = modules_by_id.get(item.module_id, "Unassigned"); module_counts[module_name] = module_counts.get(module_name, 0) + 1
        if (item.status or "").lower() not in completed_statuses: workload[item.assigned_to] = workload.get(item.assigned_to, 0) + 1
        if item in overdue: overdue_modules[module_name] = overdue_modules.get(module_name, 0) + 1
    recent = sorted(actions, key=lambda item: item.created_at or datetime.min, reverse=True)[:10]
    return {"summary": {"total_projects": len(projects), "total_clients": len({project.organization_id for project in projects}), "total_action_items": len(actions), "total_completed": len(completed), "overall_progress": round(len(completed) / len(actions) * 100) if actions else 0, "overdue": len(overdue), "blocked": len(blocked)}, "project_health": health, "module_heatmap": [{"name": name, "count": count} for name, count in module_counts.items()], "team_workload": [{"user_id": user_id, "user_name": users_by_id.get(user_id, "Unassigned"), "task_count": count} for user_id, count in sorted(workload.items(), key=lambda pair: pair[1], reverse=True)[:10]], "overdue_by_module": [{"name": name, "count": count} for name, count in overdue_modules.items()], "recent_activity": [{"id": item.id, "title": item.title, "status": item.status, "project_id": item.project_id, "created_at": (item.created_at or datetime.utcnow()).isoformat()} for item in recent]}


@app.get("/api/dashboards/master")
def master_dashboard(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role not in {"admin", "manager"}: raise HTTPException(403, "Master Control access is restricted")
    projects = db.query(Project).all(); project_ids = [project.id for project in projects]; today = date.today(); completed_statuses = {"done", "completed", "closed"}
    actions = db.query(ActionItem).filter(ActionItem.project_id.in_(project_ids)).all() if project_ids else []
    open_items = [item for item in actions if (item.status or "").lower() not in completed_statuses]; overdue = [item for item in open_items if item.eda and item.eda < today]; critical = [item for item in open_items if (item.priority or "").lower() == "critical"]
    clients = {client.id: client.name for client in db.query(Client).all()}; users = {item.id: item.name for item in db.query(User).all()}
    snapshots = []
    for project in projects:
        items = [item for item in actions if item.project_id == project.id]; done = [item for item in items if (item.status or "").lower() in completed_statuses]; late = [item for item in items if (item.status or "").lower() not in completed_statuses and item.eda and item.eda < today]; blocked = sum((item.status or "").lower() == "blocked" for item in items)
        snapshots.append({"id": project.id, "name": project.name, "client": clients.get(project.organization_id, "Unknown"), "pm": "Unassigned", "total": len(items), "completed": len(done), "progress": round(len(done) / len(items) * 100) if items else 0, "overdue": len(late), "blocked": blocked})
    workload = {}
    for item in open_items: workload[item.assigned_to] = workload.get(item.assigned_to, 0) + 1
    urgent = sorted([item for item in open_items if (item.priority or "").lower() == "critical" or (item.status or "").lower() == "blocked" or (item.eda and item.eda < today)], key=lambda item: item.eda or date.max)[:20]
    upcoming = sorted([item for item in open_items if item.eda and today <= item.eda <= today + timedelta(days=14)], key=lambda item: item.eda)[:15]
    return {"summary": {"total_projects": len(projects), "total_tasks": len(actions), "open_tasks": len(open_items), "overdue": len(overdue), "critical": len(critical)}, "projects": snapshots, "resource_allocation": [{"user_id": user_id, "user_name": users.get(user_id, "Unassigned"), "task_count": count} for user_id, count in sorted(workload.items(), key=lambda pair: pair[1], reverse=True)[:15]], "urgent_actions": [{"id": item.id, "title": item.title, "project": next((project.name for project in projects if project.id == item.project_id), "Unknown"), "status": item.status, "priority": item.priority, "eda": item.eda.isoformat() if item.eda else None} for item in urgent], "upcoming_deadlines": [{"id": item.id, "title": item.title, "project": next((project.name for project in projects if project.id == item.project_id), "Unknown"), "eda": item.eda.isoformat()} for item in upcoming]}


@app.get("/api/dashboards/qa")
def qa_dashboard(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role == "viewer": raise HTTPException(403, "QA dashboard access is restricted")
    projects = db.query(Project).all(); project_ids = [project.id for project in projects]; project_names = {project.id: project.name for project in projects}
    items = db.query(ActionItem).filter(ActionItem.project_id.in_(project_ids)).all() if project_ids else []
    def tagged(item, tag): return tag in (item.tags or [])
    testing = [item for item in items if (item.status or "").lower() == "testing"]
    failed = [item for item in items if (item.status or "").lower() == "failed" or tagged(item, "failed_test")]
    critical = [item for item in items if (item.priority or "").lower() == "critical" and tagged(item, "bug")]
    uat = [item for item in items if tagged(item, "uat_ready")]
    recent = sorted([item for item in items if item.created_at and item.created_at >= datetime.utcnow() - timedelta(days=7) and ((item.status or "").lower() in {"testing", "failed"} or tagged(item, "uat_ready"))], key=lambda item: item.created_at, reverse=True)[:15]
    def base(item): return {"id": item.id, "title": item.title, "project": project_names.get(item.project_id, "Unknown"), "priority": item.priority, "assigned_to": item.owner or "Unassigned"}
    return {"summary": {"testing_queue": len(testing), "failed_tests": len(failed), "critical_bugs": len(critical), "ready_for_uat": len(uat)}, "testing_queue": [{**base(item), "eda": item.eda.isoformat() if item.eda else None} for item in testing[:20]], "failed_tests": [{**base(item), "updated_at": (item.created_at or datetime.utcnow()).isoformat()} for item in failed[:20]], "critical_bugs": [{**base(item), "status": item.status} for item in critical[:20]], "ready_for_uat": [{"id": item.id, "title": item.title, "project": project_names.get(item.project_id, "Unknown"), "completed_at": (item.created_at or datetime.utcnow()).isoformat()} for item in uat[:20]], "recent_activity": [{"id": item.id, "title": item.title, "project": project_names.get(item.project_id, "Unknown"), "status": item.status, "updated_at": (item.created_at or datetime.utcnow()).isoformat()} for item in recent]}


def require_qa_editor(user: User):
    if user.role == "viewer": raise HTTPException(403, "QA summary editing is restricted to internal editors")


@app.get("/api/qa-summary", response_model=list[QASummaryOut])
def get_qa_summary(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_qa_editor(user)
    return db.query(QASummary).order_by(QASummary.id).all()


@app.post("/api/qa-summary", response_model=QASummaryOut, status_code=201)
def create_qa_summary(payload: QASummaryCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_qa_editor(user)
    if min(payload.total_bugs, payload.done, payload.in_progress, payload.to_do) < 0: raise HTTPException(422, "QA counts cannot be negative")
    item = QASummary(**payload.model_dump()); db.add(item); db.commit(); db.refresh(item)
    return item


@app.delete("/api/qa-summary", status_code=204)
def clear_qa_summary(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_qa_editor(user); db.query(QASummary).delete(); db.commit()


@app.get("/api/qa-summary/template")
def qa_summary_template(user: User = Depends(get_current_user)):
    require_qa_editor(user)
    return Response(content="total_bugs,done,in_progress,to_do,assignee,owner\n", media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="delivery-hub-qa-summary-template.csv"'})


@app.post("/api/qa-summary/import", response_model=list[QASummaryOut])
async def import_qa_summary(file: UploadFile = File(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_qa_editor(user)
    raw = await file.read(); filename = (file.filename or "").lower()
    try:
        if filename.endswith(".csv"):
            rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True); sheet = workbook.active
            values = list(sheet.values); headers = [str(value or "").strip() for value in (values[0] if values else [])]
            rows = [dict(zip(headers, row)) for row in values[1:]]
        else: raise ValueError("Upload a .csv or .xlsx file")
        required = {"total_bugs", "done", "in_progress", "to_do", "assignee", "owner"}
        if not rows or not required.issubset({str(key).strip().lower() for key in rows[0].keys()}): raise ValueError("Use the QA summary template columns")
        normalized = []
        for row in rows:
            item = {str(key).strip().lower(): value for key, value in row.items()}
            payload = QASummaryCreate(total_bugs=int(item.get("total_bugs") or 0), done=int(item.get("done") or 0), in_progress=int(item.get("in_progress") or 0), to_do=int(item.get("to_do") or 0), assignee=str(item.get("assignee") or ""), owner=str(item.get("owner") or ""))
            if min(payload.total_bugs, payload.done, payload.in_progress, payload.to_do) < 0: raise ValueError("QA counts cannot be negative")
            normalized.append(QASummary(**payload.model_dump()))
    except (ValueError, TypeError, OverflowError) as error:
        raise HTTPException(400, str(error))
    db.query(QASummary).delete(); db.add_all(normalized); db.commit()
    return db.query(QASummary).order_by(QASummary.id).all()


def seed(db: Session):
    existing = {client.name for client in db.query(Client).all()}
    missing = [client for client in (Client(name="Finkomm", location="Maharashtra, India"), Client(name="HHP", location="Bahrain"), Client(name="NorQ", location="Internal Operations")) if client.name not in existing]
    if missing: db.add_all(missing); db.commit()


@app.on_event("startup")
def start():
    Base.metadata.create_all(engine)
    action_columns = {column["name"] for column in inspect(engine).get_columns("action_items")}
    action_migrations = {"project_id": "INTEGER", "module_id": "INTEGER", "assigned_to": "INTEGER", "created_by": "INTEGER", "eda": "DATE", "ada": "DATE", "depends_on": "INTEGER", "tags": "JSON", "created_at": "TIMESTAMP"}
    missing = {name: sql_type for name, sql_type in action_migrations.items() if name not in action_columns}
    if missing:
        with engine.begin() as connection:
            for name, sql_type in missing.items(): connection.execute(text(f"ALTER TABLE action_items ADD COLUMN {name} {sql_type}"))
    user_columns = {column["name"] for column in inspect(engine).get_columns("users")}
    if "password_hash" not in user_columns:
        with engine.begin() as connection: connection.execute(text("ALTER TABLE users ADD COLUMN password_hash VARCHAR(255) NOT NULL DEFAULT ''"))
    if "role_id" not in user_columns:
        with engine.begin() as connection: connection.execute(text("ALTER TABLE users ADD COLUMN role_id INTEGER"))
    # Lightweight migration for workspaces created before agendas were added.
    meeting_columns = {column["name"] for column in inspect(engine).get_columns("meetings")}
    if "meetings" in inspect(engine).get_table_names() and "agenda" not in meeting_columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE meetings ADD COLUMN agenda TEXT NOT NULL DEFAULT ''"))
    if "meetings" in inspect(engine).get_table_names() and "created_at" not in meeting_columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE meetings ADD COLUMN created_at TIMESTAMP"))
    db = SessionLocal()
    try:
        seed(db)
        for role_name, description in (("admin", "Full system access"), ("manager", "Portfolio and delivery management"), ("editor", "Create and update delivery work"), ("viewer", "Read-only workspace access")):
            role_record = db.query(Role).filter(Role.name == role_name).first()
            if not role_record:
                role_record = Role(name=role_name, description=description, is_system=True, permissions=permissions_for(role_name)); db.add(role_record); db.flush()
            else:
                role_record.permissions = permissions_for(role_name)
            db.query(User).filter(User.role == role_name, User.role_id.is_(None)).update({User.role_id: role_record.id}, synchronize_session=False)
        db.commit()
        for client in db.query(Client).all():
            project = db.query(Project).filter(Project.organization_id == client.id).first()
            if not project:
                project = Project(organization_id=client.id, name=f"{client.name} Delivery"); db.add(project); db.flush()
            for module_name in ("Meetings", "Development", "Requirements"):
                if not db.query(ProjectModule).filter(ProjectModule.project_id == project.id, ProjectModule.name == module_name).first(): db.add(ProjectModule(project_id=project.id, name=module_name))
        db.commit()
    finally:
        db.close()


@app.get("/api/clients", response_model=list[ClientOut])
def clients(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    query = select(Client).order_by(Client.name)
    if user.role != "admin": query = query.where(Client.id == user.organization_id)
    return db.scalars(query).all()


@app.get("/api/projects")
def projects(client_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if client_id is not None: require_client_scope(client_id, user)
    query = select(Project).order_by(Project.name)
    if client_id is not None: query = query.where(Project.organization_id == client_id)
    elif user.role != "admin": query = query.where(Project.organization_id == user.organization_id)
    return db.scalars(query).all()


@app.get("/api/modules")
def modules(project_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    query = select(ProjectModule).order_by(ProjectModule.name)
    if project_id:
        project = scoped_record(db.get(Project, project_id), user, "Project not found")
        query = query.where(ProjectModule.project_id == project.id)
    elif user.role != "admin":
        query = query.join(Project, Project.id == ProjectModule.project_id).where(Project.organization_id == user.organization_id)
    return db.scalars(query).all()


@app.get("/api/dashboard")
def dashboard(client_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if client_id is not None: require_client_scope(client_id, user)
    if client_id is None and user.role != "admin": client_id = user.organization_id
    def filtered(model):
        statement = select(model)
        return statement.where(model.client_id == client_id) if client_id else statement
    actions = db.scalars(filtered(ActionItem)).all(); bugs = db.scalars(filtered(Bug)).all(); features = db.scalars(filtered(Feature)).all(); meetings = db.scalars(filtered(Meeting).order_by(Meeting.scheduled_for.desc())).all(); requirements = db.scalars(filtered(Requirement).order_by(Requirement.updated_at.desc())).all()
    today = date.today()
    return {"open_actions": sum(a.status != "Done" for a in actions), "overdue_actions": sum(a.status != "Done" and a.due_date is not None and a.due_date < today for a in actions), "open_bugs": sum(b.status != "Fixed" for b in bugs), "development_progress": round(sum(f.progress for f in features) / len(features)) if features else 0, "meetings": meetings, "actions": actions, "features": features, "bugs": bugs, "requirements": requirements}


@app.post("/api/meetings", response_model=MeetingOut, status_code=201)
def create_meeting(payload: MeetingCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); require_client_scope(payload.client_id, user)
    item = Meeting(**payload.model_dump()); db.add(item); db.commit(); db.refresh(item); return item


@app.patch("/api/meetings/{meeting_id}", response_model=MeetingOut)
def update_meeting(meeting_id: int, payload: MeetingUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_meeting(meeting_id, user, db)
    if not item.created_at or datetime.utcnow() - item.created_at > timedelta(hours=48): raise HTTPException(403, "Meetings can only be edited within 48 hours of creation")
    for key, value in payload.model_dump().items(): setattr(item, key, value)
    db.commit(); db.refresh(item); return item


@app.delete("/api/meetings/{meeting_id}", status_code=204)
def delete_meeting(meeting_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_delete(user); item = scoped_meeting(meeting_id, user, db)
    # Keep follow-ups in the action register, but remove their meeting link.
    db.query(ActionItem).filter(ActionItem.meeting_id == meeting_id).update({ActionItem.meeting_id: None}, synchronize_session=False)
    db.delete(item); db.commit()


@app.post("/api/meetings/{meeting_id}/generate-mom", response_model=MeetingOut)
def generate_mom(meeting_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_meeting(meeting_id, user, db)
    item.mom = f"Minutes of Meeting\n\nMeeting: {item.title}\nDate: {item.scheduled_for:%d %b %Y, %H:%M}\nAttendees: {item.attendees}\n\nDiscussion notes\n{item.notes}\n\nDecisions and action items should be reviewed and assigned by the delivery lead."
    db.commit(); db.refresh(item); return item


def scoped_meeting(meeting_id: int, user: User, db: Session) -> Meeting:
    meeting = db.get(Meeting, meeting_id)
    if not meeting or not can_access_client(user, meeting.client_id):
        raise HTTPException(404, "Meeting not found")
    return meeting


@app.get("/api/meetings/{meeting_id}/discussion-points", response_model=list[DiscussionPointOut])
def discussion_points(meeting_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    scoped_meeting(meeting_id, user, db)
    return db.query(DiscussionPoint).filter(DiscussionPoint.meeting_id == meeting_id).order_by(DiscussionPoint.created_at).all()


@app.post("/api/meetings/{meeting_id}/discussion-points", response_model=DiscussionPointOut, status_code=201)
def create_discussion_point(meeting_id: int, payload: DiscussionPointCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user)
    meeting = scoped_meeting(meeting_id, user, db)
    item = DiscussionPoint(meeting_id=meeting.id, client_id=meeting.client_id, **payload.model_dump())
    db.add(item); db.commit(); db.refresh(item); return item


@app.patch("/api/discussion-points/{discussion_id}", response_model=DiscussionPointOut)
def update_discussion_point(discussion_id: int, payload: DiscussionPointUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user)
    item = db.get(DiscussionPoint, discussion_id)
    if not item or not can_access_client(user, item.client_id): raise HTTPException(404, "Discussion point not found")
    for key, value in payload.model_dump().items(): setattr(item, key, value)
    db.commit(); db.refresh(item); return item


@app.post("/api/discussion-points/{discussion_id}/convert", response_model=ActionOut, status_code=201)
def convert_discussion_to_action(discussion_id: int, payload: ActionCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user)
    discussion = db.get(DiscussionPoint, discussion_id)
    if not discussion or discussion.client_id != user.organization_id: raise HTTPException(404, "Discussion point not found")
    values = payload.model_dump(exclude_unset=True)
    values.update({"client_id": discussion.client_id, "title": payload.title or f"From discussion: {discussion.title}", "meeting_id": discussion.meeting_id, "created_by": user.id, "status": "Open"})
    action = ActionItem(**values)
    db.add(action); db.flush()
    discussion.linked_action_item_id = action.id; discussion.status = "Closed"
    db.commit(); db.refresh(action); return action


@app.post("/api/actions", response_model=ActionOut, status_code=201)
def create_action(payload: ActionCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); require_client_scope(payload.client_id, user)
    item = ActionItem(**payload.model_dump()); db.add(item); db.commit(); db.refresh(item); return item


@app.patch("/api/actions/{action_id}/status", response_model=ActionOut)
def update_action(action_id: int, status: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_record(db.get(ActionItem, action_id), user, "Action not found")
    old_status = item.status; item.status = status; db.add(ActionItemHistory(action_item_id=item.id, field_name="status", old_value=old_status, new_value=status)); db.commit(); db.refresh(item); return item


@app.patch("/api/actions/{action_id}", response_model=ActionOut)
def edit_action(action_id: int, payload: ActionUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_record(db.get(ActionItem, action_id), user, "Action not found")
    for key, value in payload.model_dump().items():
        old_value = getattr(item, key)
        if old_value != value: db.add(ActionItemHistory(action_item_id=item.id, field_name=key, old_value=str(old_value) if old_value is not None else None, new_value=str(value) if value is not None else None))
        setattr(item, key, value)
    db.commit(); db.refresh(item); return item


@app.delete("/api/actions/{action_id}", status_code=204)
def delete_action(action_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_delete(user); item = scoped_record(db.get(ActionItem, action_id), user, "Action not found")
    db.delete(item); db.commit()


@app.post("/api/requirements", response_model=RequirementOut, status_code=201)
def create_requirement(payload: RequirementCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); require_client_scope(payload.client_id, user)
    item = Requirement(**payload.model_dump()); db.add(item); db.commit(); db.refresh(item); return item


@app.patch("/api/requirements/{requirement_id}", response_model=RequirementOut)
def update_requirement(requirement_id: int, payload: RequirementUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_record(db.get(Requirement, requirement_id), user, "Requirement not found")
    for key, value in payload.model_dump().items(): setattr(item, key, value)
    item.updated_at = datetime.utcnow(); db.commit(); db.refresh(item); return item


@app.delete("/api/requirements/{requirement_id}", status_code=204)
def delete_requirement(requirement_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_delete(user); item = scoped_record(db.get(Requirement, requirement_id), user, "Requirement not found")
    db.delete(item); db.commit()


@app.get("/api/features", response_model=list[FeatureOut])
def features(client_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if client_id is not None: require_client_scope(client_id, user)
    if client_id is None and user.role != "admin": client_id = user.organization_id
    q = select(Feature); q = q.where(Feature.client_id == client_id) if client_id is not None else q
    return db.scalars(q).all()


@app.post("/api/features", response_model=FeatureOut, status_code=201)
def create_feature(payload: FeatureCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); require_client_scope(payload.client_id, user)
    item = Feature(**payload.model_dump()); db.add(item); db.commit(); db.refresh(item); return item


@app.patch("/api/features/{feature_id}/progress", response_model=FeatureOut)
def update_feature_progress(feature_id: int, progress: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_record(db.get(Feature, feature_id), user, "Feature not found")
    item.progress = max(0, min(progress, 100)); db.commit(); db.refresh(item); return item


@app.patch("/api/features/{feature_id}", response_model=FeatureOut)
def update_feature(feature_id: int, payload: FeatureUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_record(db.get(Feature, feature_id), user, "Development item not found"); require_client_scope(payload.client_id, user)
    item.client_id = payload.client_id; item.name = payload.name; item.developer = payload.developer; item.sprint = payload.sprint; item.progress = max(0, min(payload.progress, 100)); item.target_date = payload.target_date; item.dependency = payload.dependency
    db.commit(); db.refresh(item); return item


@app.delete("/api/features/{feature_id}", status_code=204)
def delete_feature(feature_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_delete(user); item = scoped_record(db.get(Feature, feature_id), user, "Development item not found")
    db.delete(item); db.commit()


@app.get("/api/bugs", response_model=list[BugOut])
def bugs(client_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if client_id is not None: require_client_scope(client_id, user)
    if client_id is None and user.role != "admin": client_id = user.organization_id
    q = select(Bug); q = q.where(Bug.client_id == client_id) if client_id is not None else q
    return db.scalars(q).all()


@app.post("/api/bugs", response_model=BugOut, status_code=201)
def create_bug(payload: BugCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); require_client_scope(payload.client_id, user)
    item = Bug(**payload.model_dump()); db.add(item); db.commit(); db.refresh(item); return item


@app.patch("/api/bugs/{bug_id}", response_model=BugOut)
def update_bug(bug_id: int, payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user); item = scoped_record(db.get(Bug, bug_id), user, "Bug not found")
    for field in ("title", "module", "severity", "assigned_developer", "qa_owner", "status", "target_fix_date"):
        if field in payload: setattr(item, field, payload[field])
    db.commit(); db.refresh(item); return item


@app.delete("/api/bugs/{bug_id}", status_code=204)
def delete_bug(bug_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_delete(user); item = scoped_record(db.get(Bug, bug_id), user, "Bug not found")
    db.delete(item); db.commit()


@app.get("/api/bugs/import-template")
def download_bug_import_template():
    csv_template = "client_id,title,module,severity,assigned_developer,qa_owner,status,target_fix_date\n"
    return Response(content=csv_template, media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="delivery-hub-bugs-template.csv"'})


@app.post("/api/bugs/import", response_model=dict)
async def import_bugs(file: UploadFile = File(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_write(user)
    filename = (file.filename or "").lower()
    raw = await file.read()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
    elif filename.endswith((".xlsx", ".xls")):
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook.active
            values = list(sheet.values)
            if values:
                headers = [str(value or "").strip() for value in values[0]]
                rows = [dict(zip(headers, row)) for row in values[1:]]
        except ImportError:
            raise HTTPException(500, "Excel import requires openpyxl. Install backend requirements and retry.")
    else:
        raise HTTPException(400, "Upload a .csv or .xlsx file")
    aliases = {"client": "client_id", "client id": "client_id", "bug title": "title", "description": "title", "developer": "assigned_developer", "assigned developer": "assigned_developer", "qa": "qa_owner", "qa owner": "qa_owner", "fix date": "target_fix_date", "target fix date": "target_fix_date"}
    created = 0
    for raw_row in rows:
        row = {aliases.get(str(key).strip().lower(), str(key).strip().lower()): value for key, value in raw_row.items() if key is not None}
        if not row.get("title") or not row.get("client_id"): continue
        try: client_id = int(row["client_id"])
        except (TypeError, ValueError): continue
        fix_date = row.get("target_fix_date") or None
        if fix_date:
            try: fix_date = date.fromisoformat(str(fix_date)[:10])
            except ValueError: fix_date = None
        require_client_scope(client_id, user)
        item = Bug(client_id=client_id, title=str(row["title"]), module=str(row.get("module") or "General"), severity=str(row.get("severity") or "Medium"), assigned_developer=str(row.get("assigned_developer") or ""), qa_owner=str(row.get("qa_owner") or ""), status=str(row.get("status") or "Open"), target_fix_date=fix_date)
        db.add(item); created += 1
    db.commit()
    return {"created": created, "skipped": len(rows) - created}
